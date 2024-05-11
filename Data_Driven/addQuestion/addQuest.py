# Generated by Selenium IDE
import os 

import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.common.exceptions import NoSuchElementException


class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestAddQuest():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.vars = {}
    
    def teardown_method(self):
        self.driver.quit()

    def first_step(self):
        self.driver.get('https://sso.hcmut.edu.vn/cas/login?service=https%3A%2F%2Flms.hcmut.edu.vn%2Flogin%2Findex.php%3FauthCAS%3DCAS')
        time.sleep(2)
        
        self.driver.find_element(By.NAME,"username").send_keys("010552")
        self.driver.find_element(By.NAME,"password").send_keys("010552")
        self.driver.find_element(By.NAME,"submit").click()
        time.sleep(2)

        # Go to link for create new quizz
        self.driver.get('https://lms.hcmut.edu.vn/question/bank/editquestion/question.php?courseid=48411&sesskey=MORn49Z0Oy&qtype=truefalse&returnurl=%2Fquestion%2Fedit.php%3Fcourseid%3D48411%26deleteall%3D1&courseid=48411&category=63872')
        time.sleep(1)


    def test_add_quest(self, name, description, mark, expectedResult):
        self.first_step()
        
        self.driver.execute_script(f"arguments[0].setAttribute('value', '{name}')", self.driver.find_element(By.ID, 'id_name'))
        time.sleep(1)

        #Fill description
        frame = self.driver.find_element(By.ID, "id_questiontext_ifr")  # Locate the iframe element
        self.driver.switch_to.frame(frame)
        body = self.driver.find_element(By.TAG_NAME, "body")
        body.clear()
        # Fill text into the body element
        body.send_keys(description)

        self.driver.switch_to.default_content()


        self.driver.execute_script(f"arguments[0].setAttribute('value','{mark}')", self.driver.find_element(By.ID, 'id_defaultmark'))
        time.sleep(1)


        self.driver.find_element(By.ID,"id_updatebutton").click()
        time.sleep(3)

        if expectedResult == "FillName":
            error_message = self.driver.find_element(By.ID,'id_error_name').get_attribute("outerText")
            assert ("You must supply a value here" in error_message)
            
        elif expectedResult == "FillText":
            error_message = self.driver.find_element(By.ID,'id_error_questiontext').get_attribute("outerText")
            assert ("You must supply a value here" in error_message)
        
        elif expectedResult == "FillMark":
            error_message = self.driver.find_element(By.ID,'id_error_defaultmark').get_attribute("outerText")
            assert ("You must supply a value here" in error_message)

        elif expectedResult == "SaveSucess":
            title = self.driver.find_element(By.TAG_NAME,"h2").text
            assert ("Editing" in title)

        elif expectedResult == "MustPositive":
            error_message = self.driver.find_element(By.ID,'id_error_defaultmark').get_attribute("outerText")
            assert ("The default mark must be positive" in error_message)

        elif expectedResult == "MustNumber":
            error_message = self.driver.find_element(By.ID,'id_error_defaultmark').get_attribute("outerText")
            assert ("You must enter a number here" in error_message)


        time.sleep(1)
        self.driver.get('https://lms.hcmut.edu.vn/login/logout.php?sesskey=NyUKwbW7f1')
        time.sleep(1)
        self.driver.find_element(By.XPATH, "//button[text()='Continue']").click()
        time.sleep(1)



if __name__ == "__main__":
    directory = r"Data_Driven\addQuestion"
    path = os.path.join(os.getcwd(), directory) 
    os.chdir(path)
    
    excel = FileExcelReader('SecB_addQuest_data.xlsx', 'Sheet1')
    test = TestAddQuest()
    
    test.setup_method()

    nRows = excel.getRowCount()
    print(nRows)
    for row in range(2, nRows + 1):
        name = excel.readData(row,2)
        description = excel.readData(row,3)
        mark = excel.readData(row,4)
    
        expectedResult = excel.readData(row,5)
       
        if name is None:
            name = ""
        if description is None:
            description = ""
        if mark is None:
            mark = ""
    
        try:
            result = test.test_add_quest(name, description, mark, expectedResult)
            excel.writeData("Passed",row,6)
        except:
            excel.writeData("Failed",row,6)

    test.teardown_method()